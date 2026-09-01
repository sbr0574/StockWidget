"""拉取全市场代码并写入仓库 resources/ 下的 JSON。

脚本可由 Azure Ubuntu 定时任务或 GitHub Actions 调用，保持自包含，
不依赖 StockWidget 运行时模块。沪深股票和基金、港股股票和基金、美股上市
证券直接使用交易所官方接口，其余市场沿用东财或新浪接口。
"""
import csv
import json
import os
import re
import sys
import time
import traceback
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone
from io import BytesIO, StringIO
from threading import Lock

import pandas as pd
import requests

from openpyxl import load_workbook
from pypinyin import Style, pinyin

# 脚本位于 <root>/scripts/。服务器可用 CODES_OUTPUT_DIR 指向 codes-data
# 工作树，确保失败分类保留数据分支中的上一版文件。
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.environ.get("CODES_OUTPUT_DIR", os.path.join(ROOT, "resources"))

STATUS_FILE = "codes_update_status.json"
US_CN_ALIAS_CACHE_FILE = "cache_us_cn_aliases.json"
TASK_ATTEMPTS = 3
TASK_RETRY_SECONDS = (2, 5)

_EM_CLIST_URL = "https://push2delay.eastmoney.com/api/qt/clist/get"
_USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
_EM_HEADERS = {
    "User-Agent": _USER_AGENT,
    "Referer": "https://quote.eastmoney.com/center/gridlist.html",
}
_HKEX_SECURITIES_EN_URL = (
    "https://www.hkex.com.hk/eng/services/trading/securities/"
    "securitieslists/ListOfSecurities.xlsx"
)
_HKEX_SECURITIES_ZH_URL = (
    "https://www.hkex.com.hk/chi/services/trading/securities/"
    "securitieslists/ListOfSecurities_c.xlsx"
)
_HKEX_HEADERS = {
    "User-Agent": _USER_AGENT,
    "Referer": "https://www.hkex.com.hk/Services/Trading/Securities/Securities-Lists",
}
_HKEX_CATEGORY_TYPES = {
    "Equity": "港",
    "Exchange Traded Products": "基",
    "Real Estate Investment Trusts": "基",
}
_NASDAQ_LISTED_URL = "https://www.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt"
_NASDAQ_OTHER_LISTED_URL = (
    "https://www.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt"
)
_NASDAQ_HEADERS = {
    "User-Agent": _USER_AGENT,
    "Referer": "https://www.nasdaqtrader.com/trader.aspx?id=symboldirdefs",
}
_SINA_HEADERS = {
    "User-Agent": _USER_AGENT,
    "Referer": "https://vip.stock.finance.sina.com.cn/",
}
_DF_COLUMNS = ["code", "name", "name_en", "type", "market"]
_SZSE_REQUEST_LOCK = Lock()


# ----------------- 工具函数 -----------------

def _has_cjk(text: str) -> bool:
    """判断字符串是否包含 CJK 中日韩字符。"""
    return any("\u4e00" <= ch <= "\u9fff" for ch in str(text or ""))


def _concat_dedup(frames: list[pd.DataFrame]) -> pd.DataFrame:
    df = pd.concat(frames, ignore_index=True)
    return df.drop_duplicates(subset=["market", "code"]).reset_index(drop=True)


# ----------------- 交易所数据 -----------------

def _exchange_code(value) -> str:
    """把交易所工作簿中的数字代码规范为六位字符串。"""
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return ""
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return text.split(".", 1)[0].zfill(6)
    return ""


def _hkex_code(value) -> str:
    """把港交所工作簿中的证券代码规范为五位字符串。"""
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return ""
    if re.fullmatch(r"\d{1,5}(?:\.0+)?", text):
        return text.split(".", 1)[0].zfill(5)
    return ""


def _rows_frame(rows: list[tuple[str, str, str, str, str]]) -> pd.DataFrame:
    return pd.DataFrame(rows, columns=_DF_COLUMNS).drop_duplicates(
        subset=["market", "code"]
    ).reset_index(drop=True)

def _stock_sh_name_code(symbol: str) -> pd.DataFrame:
    """获取上交所股票列表。"""
    symbol_map = {
        "主板A股": ("1", "A_STOCK_CODE"),
        "主板B股": ("2", "B_STOCK_CODE"),
        "科创板": ("8", "A_STOCK_CODE"),
    }
    stock_type, code_field = symbol_map[symbol]
    response = requests.get(
        "https://query.sse.com.cn/sseQuery/commonQuery.do",
        params={
            "STOCK_TYPE": stock_type,
            "REG_PROVINCE": "",
            "CSRC_CODE": "",
            "STOCK_CODE": "",
            "sqlId": "COMMON_SSE_CP_GPJCTPZ_GPLB_GP_L",
            "COMPANY_STATUS": "2,4,5,7,8",
            "type": "inParams",
            "isPagination": "true",
            "pageHelp.cacheSize": "1",
            "pageHelp.beginPage": "1",
            "pageHelp.pageSize": "10000",
            "pageHelp.pageNo": "1",
            "pageHelp.endPage": "1",
        },
        headers={
            "Host": "query.sse.com.cn",
            "Pragma": "no-cache",
            "Referer": "https://www.sse.com.cn/assortment/stock/list/share/",
            "User-Agent": _USER_AGENT,
        },
        timeout=15,
    )
    response.raise_for_status()
    rows = []
    for item in (response.json() or {}).get("result") or []:
        code = _exchange_code(item.get(code_field))
        name = str(item.get("SEC_NAME_CN") or "").strip()
        mtype = "沪" if stock_type in {"1", "2"} else "科"
        if code and name:
            rows.append((code, name, "", mtype, "sh"))
    return _rows_frame(rows)

def _stock_sh_all() -> pd.DataFrame:
    """顺序获取上交所 A 股、B 股和科创板。"""
    frames = []
    for label, symbol in (
        ("沪A", "主板A股"),
        ("沪B", "主板B股"),
        ("科创板", "科创板"),
    ):
        frame = _stock_sh_name_code(symbol)
        print(f"{label}: {len(frame)} 条", flush=True)
        if frame.empty:
            raise ValueError(f"{label}返回列表为空")
        frames.append(frame)
    return _concat_dedup(frames)

def _szse_xlsx(catalog_id: str, tab_key: str, referer: str) -> pd.DataFrame:
    """下载并解析一次深交所 XLSX，失败后由分类任务统一重试。"""
    url = (
        "https://www.szse.cn/api/report/ShowReport"
        if catalog_id == "1110"
        else "https://fund.szse.cn/api/report/ShowReport"
    )
    with _SZSE_REQUEST_LOCK:
        response = requests.get(
            url,
            params={
                "SHOWTYPE": "xlsx",
                "CATALOGID": catalog_id,
                "TABKEY": tab_key,
                "random": f"{time.time():.16f}",
            },
            headers={
                "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                "Connection": "close",
                "Referer": referer,
                "User-Agent": _USER_AGENT,
            },
            timeout=(15, 30),
        )
        response.raise_for_status()
        if not response.content.startswith(b"PK"):
            content_type = response.headers.get("Content-Type", "unknown")
            raise ValueError(f"深交所返回的不是 XLSX: {content_type}")
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            table = pd.read_excel(
                BytesIO(response.content), engine="openpyxl", dtype=str
            )
        if table.empty:
            raise ValueError("深交所 XLSX 内容为空")
        return table

def _stock_sz_a_name_code() -> pd.DataFrame:
    """下载深交所 A 股表，并按交易所板块字段拆分主板和创业板。"""
    table = _szse_xlsx(
        "1110", "tab1", "https://www.szse.cn/market/product/stock/list/index.html"
    )
    frames = []
    for board in ("主板", "创业板"):
        rows = []
        board_rows = table[table["板块"].astype(str).str.strip() == board]
        for _, item in board_rows.iterrows():
            code = _exchange_code(item.get("A股代码"))
            name = str(item.get("A股简称") or "").strip()
            mtype = "深" if board == "主板" else "创"
            if code and name:
                rows.append((code, name, "", mtype, "sz"))
        frame = _rows_frame(rows)
        print(f"深证{board}: {len(frame)} 条", flush=True)
        if frame.empty:
            raise ValueError(f"深证{board}返回列表为空")
        frames.append(frame)
    return _concat_dedup(frames)

def _stock_sz_b_name_code() -> pd.DataFrame:
    """获取深交所 B 股列表。"""
    table = _szse_xlsx(
        "1110", "tab2", "https://www.szse.cn/market/product/stock/list/index.html"
    )
    rows = []
    for _, item in table.iterrows():
        code = _exchange_code(item.get("B股代码"))
        name = str(item.get("B股简称") or "").strip()
        if code and name:
            rows.append((code, name, "", "深", "sz"))
    print(f"深B: {len(rows)} 条", flush=True)
    if rows == []:
        raise ValueError("深B返回列表为空")
    return _rows_frame(rows)

def _stock_sz_all() -> pd.DataFrame:
    """顺序获取深交所 A 股、创业板和 B 股。"""
    a_frame = _stock_sz_a_name_code()
    b_frame = _stock_sz_b_name_code()
    return _concat_dedup([a_frame, b_frame])

def _fund_sse_rows() -> dict[str, list[tuple[str, str, str, str, str]]]:
    """获取上交所上市交易基金全表并按官网 fundType 分类。"""
    result = {"ETF基金": [], "LOF基金": [], "封闭式基金": []}
    page = 1
    while True:
        response = requests.get(
            "https://query.sse.com.cn/commonSoaQuery.do",
            params={
                "isPagination": "true",
                "pageHelp.pageSize": "2000",
                "pageHelp.pageNo": str(page),
                "pageHelp.beginPage": str(page),
                "pageHelp.cacheSize": "1",
                "pageHelp.endPage": str(page),
                "pagecache": "false",
                "sqlId": "FUND_LIST",
                "order": "fundCode|asc",
            },
            headers={
                "Referer": "https://www.sse.com.cn/assortment/fund/list/",
                "User-Agent": _USER_AGENT,
            },
            timeout=15,
        )
        response.raise_for_status()
        payload = response.json() or {}
        for item in payload.get("result") or []:
            fund_type = str(item.get("fundType") or "").strip()
            if fund_type == "00":
                category = "ETF基金"
            elif fund_type == "10":
                category = "LOF基金"
            elif fund_type in {"40", "50"}:
                category = "封闭式基金"
            else:
                continue
            code = _exchange_code(item.get("fundCode"))
            name = str(item.get("secNameFull") or item.get("fundAbbr") or "").strip()
            if code and name:
                result[category].append((code, name, "", "基", "sh"))
        page_help = payload.get("pageHelp") or {}
        if page >= int(page_help.get("pageCount") or 1):
            break
        page += 1
    return result

def _fund_szse_rows() -> dict[str, list[tuple[str, str, str, str, str]]]:
    """获取深交所基金全表，并使用工作簿的基金类别字段分类。"""
    table = _szse_xlsx(
        "1000_lf",
        "tab1",
        "https://fund.szse.cn/marketdata/fundslist/index.html",
    )
    category_map = {
        "ETF": "ETF基金",
        "LOF": "LOF基金",
        "不动产基金": "封闭式基金",
        "基础设施基金": "封闭式基金",
        "REITs": "封闭式基金",
        "REITS": "封闭式基金",
    }
    result = {"ETF基金": [], "LOF基金": [], "封闭式基金": []}
    for _, item in table.iterrows():
        category = category_map.get(str(item.get("基金类别") or "").strip())
        code = _exchange_code(item.get("基金代码"))
        name = str(item.get("基金简称") or "").strip()
        if category and code and name:
            result[category].append((code, name, "", "基", "sz"))
    return result

def _fund_exchange_all() -> pd.DataFrame:
    """合并沪深交易所官方基金列表"""
    sh_rows = _fund_sse_rows()
    sz_rows = _fund_szse_rows()
    frames = []
    for category in ("ETF基金", "LOF基金", "封闭式基金"):
        frame = _rows_frame(sh_rows[category] + sz_rows[category])
        print(f"{category}: {len(frame)} 条", flush=True)
        frames.append(frame)
    return _concat_dedup(frames)


# ----------------- 港交所数据 -----------------

def _hkex_xlsx_records(url: str, required_columns: tuple[str, ...]) -> list[dict]:
    """下载港交所完整证券名录，并返回指定列。"""
    response = requests.get(
        url,
        headers=_HKEX_HEADERS,
        timeout=(15, 60),
    )
    response.raise_for_status()
    if not response.content.startswith(b"PK"):
        content_type = response.headers.get("Content-Type", "unknown")
        raise ValueError(f"港交所返回的不是 XLSX: {content_type}")

    workbook = load_workbook(
        BytesIO(response.content), read_only=True, data_only=True
    )
    try:
        if "ListOfSecurities" not in workbook.sheetnames:
            raise ValueError("港交所工作簿缺少 ListOfSecurities 工作表")
        sheet = workbook["ListOfSecurities"]
        # 港交所文件的 dimension 当前仅声明到第 8 行；只读模式必须重置，
        # 否则 openpyxl 会忽略后续一万余条证券记录。
        sheet.reset_dimensions()
        row_iter = sheet.iter_rows(values_only=True)
        column_indexes = None
        for values in row_iter:
            cells = [str(value or "").strip() for value in values]
            if all(column in cells for column in required_columns):
                column_indexes = {
                    column: cells.index(column) for column in required_columns
                }
                break
        if column_indexes is None:
            raise ValueError(
                "港交所工作簿缺少必要列: " + ", ".join(required_columns)
            )

        records = []
        for values in row_iter:
            record = {
                column: values[index] if index < len(values) else None
                for column, index in column_indexes.items()
            }
            if any(value not in (None, "") for value in record.values()):
                records.append(record)
        return records
    finally:
        workbook.close()


def _stock_hk_name_code() -> pd.DataFrame:
    """从港交所官方名录获取港股股票、交易所买卖产品及 REIT。"""
    english_records = _hkex_xlsx_records(
        _HKEX_SECURITIES_EN_URL,
        ("Stock Code", "Name of Securities", "Category"),
    )
    chinese_records = _hkex_xlsx_records(
        _HKEX_SECURITIES_ZH_URL,
        ("股份代號", "股份名稱"),
    )
    chinese_names = {
        code: str(item.get("股份名稱") or "").strip()
        for item in chinese_records
        if (code := _hkex_code(item.get("股份代號")))
    }

    rows = []
    stock_count = 0
    fund_count = 0
    for item in english_records:
        category = str(item.get("Category") or "").strip()
        mtype = _HKEX_CATEGORY_TYPES.get(category)
        if not mtype:
            continue
        code = _hkex_code(item.get("Stock Code"))
        name_en = str(item.get("Name of Securities") or "").strip()
        name = chinese_names.get(code, "")
        if not code or not (name or name_en):
            continue
        rows.append((code, name, name_en, mtype, "hk"))
        if mtype == "基":
            fund_count += 1
        else:
            stock_count += 1

    print(f"港股股票: {stock_count} 条", flush=True)
    print(f"港股基金: {fund_count} 条", flush=True)
    return _rows_frame(rows)


# ----------------- 东财数据 -----------------

def _em_clist_all(fs: str, fid: str = "f12", fields: str = "f12,f14") -> list[dict]:
    """东财 clist 分页拉全市场列表；分页失败由分类任务统一重试。"""
    _em_page_size = 500
    rows = []
    page = 1
    total = None
    while True:
        r = requests.get(
            _EM_CLIST_URL,
            params={
                "pn": page, "pz": _em_page_size, "po": 1, "np": 1, "fltt": 2, "invt": 2,
                "ut": "bd1d9ddb04089700cf9c27f6f7426281",
                "fid": fid, "fs": fs, "fields": fields,
            },
            headers=_EM_HEADERS,
            timeout=10,
        )
        r.raise_for_status()
        data = ((r.json() or {}).get("data") or {})
        diff = data.get("diff") or []
        if not isinstance(diff, list):
            raise ValueError("东财代码列表返回格式错误")
        if data.get("total") is not None:
            total = int(data.get("total") or 0)
        if not diff:
            if total is not None and len(rows) < total:
                raise ValueError(f"东财代码列表提前结束: {len(rows)}/{total}")
            break
        rows.extend(diff)
        if total is not None and len(rows) >= total:
            break
        if total is None and len(diff) < _em_page_size:
            break
        page += 1
    return rows

def _em_stock_df(fs: str, mtype: str, market: str) -> pd.DataFrame:
    """从东财 clist 拉取一个分类的股票/基金/指数列表 返回df"""
    rows = []
    for d in _em_clist_all(fs, fields="f12,f14"):
        code = str(d.get("f12") or "").strip()
        name = str(d.get("f14") or "").strip()
        if not code or name.startswith("退市") or name.endswith("退"):
            continue
        rows.append((code, name, "", mtype, market))
    df = pd.DataFrame(rows, columns=_DF_COLUMNS)
    return df.drop_duplicates(subset=["market", "code"]).reset_index(drop=True)

def _index_cn_em() -> pd.DataFrame:
    """东财沪深指数列表。"""
    return _concat_dedup([
        _em_stock_df("m:1+t:1", "指", market="sh"),
        _em_stock_df("m:0+t:5", "指", market="sz"),
    ])


def _clean_us_cn_aliases(value) -> dict[str, str]:
    """清理别名缓存，只保留带中文字符的 code -> name。"""
    aliases = {}
    for raw_code, raw_name in (value if isinstance(value, dict) else {}).items():
        code = str(raw_code or "").strip().lower()
        name = str(raw_name or "").strip()
        if code and _has_cjk(name):
            aliases[code] = name
    return aliases


def _fetch_us_cn_aliases() -> dict[str, str]:
    """从东财拉取美股中文展示名；不参与官方证券范围判断。"""
    aliases = {}
    for item in _em_clist_all("m:105,m:106,m:107", fid="f20"):
        code = str(item.get("f12") or "").strip().lower()
        name = str(item.get("f14") or "").strip()
        if code and _has_cjk(name):
            aliases[code] = name
    if not aliases:
        raise ValueError("东财美股列表未返回中文别名")
    return aliases


def _load_us_cn_alias_state(cache_path: str) -> tuple[dict, dict[str, str]]:
    """读取由数据分支管理、但不向客户端下发的中文别名表。"""
    cache = _load_json(cache_path)
    return cache, _clean_us_cn_aliases(cache.get("aliases"))


def _stock_us_name_code(
    today: date | None = None,
    output_dir: str | None = None,
    cache_path: str | None = None,
) -> pd.DataFrame:
    """每日以 Nasdaq 官方目录为准；跨月后的首次运行从东财更新中文别名。"""
    output_dir = output_dir or OUTPUT_DIR
    cache_path = cache_path or os.path.join(output_dir, US_CN_ALIAS_CACHE_FILE)
    today = today or datetime.now(timezone(timedelta(hours=8))).date()

    official_rows = []
    rows = []
    sources = (
        (_NASDAQ_LISTED_URL, "Symbol"),
        # CQS Symbol 最接近现有行情接口使用的代码：类别股用点号、
        # 优先股用小写 p、权证用 .WS；下方统一转换为下划线形式。
        (_NASDAQ_OTHER_LISTED_URL, "CQS Symbol"),
    )
    for url, symbol_field in sources:
        response = requests.get(
            url,
            headers=_NASDAQ_HEADERS,
            timeout=(15, 60),
        )
        response.raise_for_status()
        text = response.content.decode("utf-8-sig")
        for item in csv.DictReader(StringIO(text), delimiter="|"):
            if str(item.get("Test Issue") or "").strip().upper() != "N":
                continue
            symbol = str(item.get(symbol_field) or "").strip()
            name = str(item.get("Security Name") or "").strip()
            if not symbol or not name or symbol.startswith("File Creation Time"):
                continue
            code = symbol.replace(".", "_").replace("p", "_").lower()
            official_rows.append((code, name))

    official_codes = {code for code, _ in official_rows}
    cache, aliases = _load_us_cn_alias_state(cache_path)
    current_month = today.strftime("%Y-%m")
    cached_month = str(cache.get("last_update") or "")[:7]
    if cached_month != current_month or not aliases:
        try:
            refreshed = _fetch_us_cn_aliases()
            aliases = {
                code: name for code, name in refreshed.items()
                if code in official_codes
            }
            if not aliases:
                raise ValueError("东财中文别名与 Nasdaq 官方代码无匹配")
            _save_json(
                cache_path,
                {"last_update": today.isoformat(), "aliases": aliases},
            )
            print(f"美股中文别名: 东财更新 {len(aliases)} 条", flush=True)
        except Exception as exc:
            print(
                f"::warning::美股中文别名更新失败，继续使用缓存: {exc}",
                flush=True,
            )
    else:
        aliases = {
            code: name for code, name in aliases.items()
            if code in official_codes
        }
        print(f"美股中文别名: 使用缓存 {len(aliases)} 条", flush=True)

    for code, official_name in official_rows:
        rows.append(
            (code, aliases.get(code, official_name), official_name, "美", "us")
        )
    return _rows_frame(rows)


# ----------------- 离线数据 -----------------

def _stock_hk_index_name_code() -> pd.DataFrame:
    """港股指数列表。"""
    _HK_INDEXES = (
        ("ces100", "中华港股通精选100指数"),
        ("ces120", "中华120指数"),
        ("ces280", "中华280指数"),
        ("ces300", "中华沪深港300指数"),
        ("cesa80", "中华A80指数"),
        ("cesg10", "中华博彩业指数"),
        ("ceshkm", "中华香港内地指数"),
        ("cscmc", "中证内地消费指数"),
        ("cshk100", "中证香港100指数"),
        ("cshkdiv", "中证香港红利港币指数"),
        ("cshklc", "中证香港上市可交易内地消费指数"),
        ("cshklre", "中证香港上市可交易内地地産指数"),
        ("cshkmcs", "中证香港中盘精选港币指数"),
        ("cshkme", "中证香港内地股港元指数"),
        ("cshkpe", "中证香港内地民营企业指数"),
        ("cshkse", "中证香港内地国有企业指数"),
        ("csrhk50", "中证锐联香港基本面50港币指数"),
        ("gem", "标普香港创业板指数"),
        ("hkl", "标普香港大型股指数"),
        ("hscci", "恒生香港中资企业指数"),
        ("hscei", "恒生中国企业指数"),
        ("hsi", "恒生指数"),
        ("hsmbi", "恒生中国内地银行指数"),
        ("hsmogi", "恒生中国内地石油及天然气指数"),
        ("hsmpi", "恒生中国内地地产指数"),
        ("hstech", "恒生科技指数"),
        ("vhsi", "恒指波幅指数"),
    )
    rows = [(code, name, "", "指", "hk") for code, name in _HK_INDEXES]
    return pd.DataFrame(rows, columns=_DF_COLUMNS)

def _stock_global_index_name_code() -> pd.DataFrame:
    """全球主要指数列表"""
    _GLOBAL_INDEXES = (
        ("aex", "荷兰AEX综合指数"),
        ("as51", "澳大利亚标准普尔200指数"),
        ("cac", "法CAC40指数"),
        ("case", "埃及CASE 30指数"),
        ("dax", "德国DAX 30种股价指数"),
        ("ftsemib", "富时意大利MIB指数"),
        ("gsptse", "加拿大S&P/TSX综合指数"),
        ("ibex", "西班牙IBEX指数"),
        ("ibov", "巴西BOVESPA股票指数"),
        ("indexcf", "俄罗斯MICEX指数"),
        ("jci", "印度尼西亚雅加达综合指数"),
        ("kospi", "首尔综合指数"),
        ("mxx", "墨西哥BOLSA指数"),
        ("nky", "日经225指数"),
        ("nz250", "新西兰NZSE 50指数"),
        ("sensex", "印度孟买SENSEX指数"),
        ("swi20", "瑞士股票指数"),
        ("sx5e", "欧洲Stoxx50指数"),
        ("twjq", "中国台湾加权指数"),
        ("ukx", "英国富时100指数"),
    )
    rows = [(code, name, "", "指", "gb") for code, name in _GLOBAL_INDEXES]
    return pd.DataFrame(rows, columns=_DF_COLUMNS)

def _stock_us_index_name_code() -> pd.DataFrame:
    """美股主要指数"""
    rows = [
        ("ixic", "纳斯达克综合指数", "", "指", "us"),
        ("dji", "道琼斯工业平均指数", "", "指", "us"),
        ("inx", "标普500指数", "", "指", "us"),
        ("ndx", "纳斯达克100指数", "", "指", "us"),
    ]
    return pd.DataFrame(rows, columns=_DF_COLUMNS)

def _offline_index_name_code() -> pd.DataFrame:
    """合并无需联网更新的港股、美股和全球主要指数。"""
    frames = []
    for label, func in (
        ("港股指数", _stock_hk_index_name_code),
        ("美股指数", _stock_us_index_name_code),
        ("全球股指", _stock_global_index_name_code),
    ):
        frame = func()
        print(f"{label}: {len(frame)} 条", flush=True)
        frames.append(frame)
    return _concat_dedup(frames)


# ----------------- 股指列表 -----------------

def _tasks() -> list[tuple]:
    """每个任务负责一个 JSON；耗时较长的任务优先提交到线程池。"""
    return [
        ("stock_us.json", "美股", _stock_us_name_code, (), {}),
        ("stock_hk.json", "港股", _stock_hk_name_code, (), {}),
        ("futures_sh.json", "上期所期货", futures_info_all, (), {}),
        ("stock_sh.json", "沪市股票", _stock_sh_all, (), {}),
        ("stock_sz.json", "深市股票", _stock_sz_all, (), {}),
        ("fund_cn.json", "沪深基金", _fund_exchange_all, (), {}),
        ("stock_bj.json", "京市", _em_stock_df, ("m:0+t:81+s:2048", "京"), {"market": "bj"}),
        ("index_cn.json", "国内指数", _index_cn_em, (), {}),
        ("index_global.json", "全球指数", _offline_index_name_code, (), {}),
    ]


def _run_task(task: tuple) -> tuple[pd.DataFrame | None, str | None]:
    """运行单个分类；首次失败后再重试两次，不影响其他分类。"""
    _, label, func, args, kwargs = task
    last_error = None
    for attempt in range(1, TASK_ATTEMPTS + 1):
        try:
            result = func(*args, **kwargs)
            if result.empty:
                raise ValueError("返回列表为空")
            print(f"{label}更新完成: {len(result)} 条", flush=True)
            return result, None
        except Exception as exc:
            last_error = exc
            traceback.print_exc()
            if attempt < TASK_ATTEMPTS:
                delay = TASK_RETRY_SECONDS[attempt - 1]
                print(
                    f"{label}第 {attempt} 次拉取失败，{delay} 秒后重试: {exc}",
                    flush=True,
                )
                time.sleep(delay)
    message = str(last_error or "未知错误")
    print(f"::warning::{label}连续 {TASK_ATTEMPTS} 次拉取失败，保留旧文件: {message}")
    return None, message


def _run_tasks(tasks: list[tuple]) -> dict[str, tuple[pd.DataFrame | None, str | None]]:
    """并发执行分类任务，返回每个文件的最终结果。"""
    results = {}
    with ThreadPoolExecutor(max_workers=min(6, len(tasks))) as pool:
        pending = {pool.submit(_run_task, task): task for task in tasks}
        for future in as_completed(pending):
            filename = pending[future][0]
            try:
                results[filename] = future.result()
            except Exception as exc:
                traceback.print_exc()
                results[filename] = (None, str(exc))
    return results


# ----------------- 上期所期货 -----------------

def _futures_nodes() -> list[str]:
    """从新浪节点脚本中提取上期所及上期能源节点名。"""
    url = "https://vip.stock.finance.sina.com.cn/quotes_service/view/js/qihuohangqing.js"
    r = requests.get(url, headers=_SINA_HEADERS, timeout=15)
    r.raise_for_status()
    r.encoding = "gb2312"
    match = re.search(r"\bshfe\s*:\s*\[(.*?)\]\s*,\s*cffex\s*:", r.text, re.DOTALL)
    if not match:
        raise ValueError("新浪期货节点脚本中缺少 shfe 列表")
    return re.findall(r"\[\s*'[^']*'\s*,\s*'([^']+)'", match.group(1))

def _stock_shfe_futures() -> pd.DataFrame:
    """上期所全部期货合约（含上期能源；新浪 getHQFuturesData 按品种遍历）。"""
    rows = []
    for node in _futures_nodes():
        for page in range(1, 4):
            r = requests.get(
                "https://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/Market_Center.getHQFuturesData",
                params={"page": page, "sort": "position", "asc": "0", "node": node, "base": "futures"},
                headers=_SINA_HEADERS,
                timeout=8,
            )
            r.raise_for_status()
            contracts = r.json()
            if not isinstance(contracts, list):
                raise ValueError("新浪期货接口返回格式错误")
            if not contracts:
                break
            for contract in contracts:
                symbol = str(contract.get("symbol", "") or "").strip().lower()
                name = str(contract.get("name", "") or "").strip()
                if symbol:
                    rows.append((symbol, name))
            if len(contracts) < 20:
                break
    seen, uniq = set(), []
    for c, n in rows:
        if c not in seen:
            seen.add(c)
            uniq.append((c, n))
    rows = [(code, name, "", "期", "") for code, name in uniq]
    return pd.DataFrame(rows, columns=_DF_COLUMNS)

def futures_info_all() -> pd.DataFrame:
    """上期所期货"""
    df = _stock_shfe_futures()
    print(f"期货: {len(df)} 条", flush=True)
    return df


# ---------------- 市场列表字典生成 ----------------

def _to_halfwidth(text: str) -> str:
    """全角字符转半角, 避免全角字母匹配不上搜索。"""
    out = []
    for ch in str(text or ""):
        o = ord(ch)
        if o == 0x3000:
            out.append(" ")                      # 全角空格
        elif 0xFF01 <= o <= 0xFF5E:
            out.append(chr(o - 0xFEE0))          # 全角 -> 半角
        else:
            out.append(ch)
    return "".join(out)

def _name_pinyin(name: str) -> tuple[str, str]:
    """中文名转拼音全拼和首字母缩写, 非中文返回空串"""
    text = _to_halfwidth(name).strip()
    if not text:
        return "", ""
    text = text.replace(" ", "")
    py_full = "".join(x[0] for x in pinyin(text, style=Style.NORMAL, strict=False))
    py_abbr = "".join(x[0] for x in pinyin(text, style=Style.FIRST_LETTER, strict=False))
    return py_full.lower(), py_abbr.lower()

def _normalize_code(code: str, market: str) -> str:
    """标准化证券代码：沪深京 6 位、港股 5 位、其他原样返回"""
    code = str(code or "").strip().lower()
    market = str(market or "").strip().lower()
    if market in {"sh", "sz", "bj"} and code.isdigit():
        return code.zfill(6)
    if market == "hk" and code.isdigit():
        return code.zfill(5)
    return code

def _code_sort_key(item: tuple[str, dict]) -> tuple:
    """按市场顺序、市场、代码排序"""
    _market_order = {"sh": 0, "sz": 1, "bj": 2, "hk": 3, "us": 4, "gb": 5, "": 6}
    key, entry = item
    market = str(entry.get("market", "") or "").strip().lower()
    code = str(entry.get("code", "") or "").strip().lower()
    return (_market_order.get(market, 99), market, code, key)

def _df_to_dict(df: pd.DataFrame) -> dict:
    """把 [code,name,name_en,type,market] 的 df 转成有序字典。"""
    codes = {}
    for _, row in df.iterrows():
        market = str(row.get("market", "") or "").strip().lower()
        code = _normalize_code(row.get("code", ""), market)
        if not code:
            continue
        name = _to_halfwidth(str(row.get("name", "") or "")).strip()
        name_en = _to_halfwidth(str(row.get("name_en", "") or "")).strip()
        if not name and name_en: name = name_en
        mtype = str(row.get("type", "") or "").strip()
        py_full, py_abbr = _name_pinyin(name) if _has_cjk(name) else ("", "")
        entry = {
            "code": code,
            "market": market,
            "type": mtype,
            "name": name,
            "name_en": name_en,
            "py": py_full,
            "abbr": py_abbr,
        }
        codes[market + code] = entry
    return dict(sorted(codes.items(), key=_code_sort_key))


# ---------------- 拉取市场列表并保存 ----------------

def _load_json(path: str) -> dict:
    try:
        with open(path, "r", encoding="utf-8") as file:
            data = json.load(file)
        return data if isinstance(data, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def _save_json(path: str, data: dict) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)
        file.write("\n")
    os.replace(tmp, path)


def _iso_now() -> str:
    return datetime.now(timezone(timedelta(hours=8))).isoformat(timespec="seconds")


def _file_status_on_failure(previous: dict, existing: dict) -> dict:
    return {
        "last_checked": str(previous.get("last_checked") or ""),
        "last_update": str(
            existing.get("last_update") or previous.get("last_update") or ""
        ),
        "updated": False,
        "error": True,
    }


def update_code_files(output_dir: str = OUTPUT_DIR) -> dict:
    """更新所有分类并写状态文件；某一分类失败时保留其现有 JSON。"""
    started_at = _iso_now()
    run_date = started_at[:10]
    status_path = os.path.join(output_dir, STATUS_FILE)
    previous_status = _load_json(status_path).get("files", {})
    tasks = _tasks()
    results = _run_tasks(tasks)
    file_states = {}

    for filename, _, _, _, _ in tasks:
        path = os.path.join(output_dir, filename)
        existing = _load_json(path)
        frame, error = results.get(filename, (None, "任务未返回结果"))
        previous = previous_status.get(filename, {})
        if frame is None or error:
            file_states[filename] = _file_status_on_failure(previous, existing)
            continue

        codes = _df_to_dict(frame)
        if not codes:
            file_states[filename] = _file_status_on_failure(previous, existing)
            continue

        changed = existing.get("codes") != codes or not existing.get("last_update")
        last_update = run_date if changed else str(existing.get("last_update") or run_date)
        if changed:
            _save_json(path, {"last_update": last_update, "codes": codes})
        file_states[filename] = {
            "last_checked": run_date,
            "last_update": last_update,
            "updated": changed,
            "error": False,
        }
        action = "已更新" if changed else "无变化"
        print(f"{filename}: {len(codes)} 条，{action}", flush=True)

    status = {
        "run_date": run_date,
        "started_at": started_at,
        "completed_at": _iso_now(),
        "files": file_states,
    }
    _save_json(status_path, status)
    failed = [name for name, state in file_states.items() if state["error"]]
    if failed:
        print(f"本次失败分类: {', '.join(failed)}；其他文件已正常处理", flush=True)
    return status


def main() -> int:
    try:
        update_code_files()
        return 0
    except Exception as exc:
        traceback.print_exc()
        print(f"::error::代码列表更新流程失败: {exc}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
